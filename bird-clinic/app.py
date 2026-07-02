import os
import json
import jdatetime
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for, flash,
                   send_file, send_from_directory)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import io

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'bird-clinic-secret-key-2025')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///bird_clinic.db')
# Fix for Render/Neon: postgres:// → postgresql://
if app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
    'pool_size': 5,
    'max_overflow': 0,
}
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'لطفاً وارد شوید'

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

IRAN_TZ = timezone(timedelta(hours=3, minutes=30))

def shamsi_now():
    now = datetime.now(IRAN_TZ)
    return jdatetime.datetime.fromgregorian(datetime=now).strftime('%Y/%m/%d')

def shamsi_now_full():
    now = datetime.now(IRAN_TZ)
    return jdatetime.datetime.fromgregorian(datetime=now).strftime('%Y/%m/%d - %H:%M')

def shamsi_time():
    return datetime.now(IRAN_TZ).strftime('%H:%M')

def shamsi_today_date():
    return jdatetime.date.today()

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('دسترسی محدود به ادمین', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# ══════════════════════════════════════
# DRUG DATABASE
# ══════════════════════════════════════
VET_DRUGS = [
    {"fa": "انروفلوکساسین", "en": "Enrofloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "آموکسی‌سیلین", "en": "Amoxicillin", "type": "آنتی‌بیوتیک"},
    {"fa": "داکسی‌سایکلین", "en": "Doxycycline", "type": "آنتی‌بیوتیک"},
    {"fa": "تایلوزین", "en": "Tylosin", "type": "آنتی‌بیوتیک"},
    {"fa": "لینکواسپکتین", "en": "Lincospectin", "type": "آنتی‌بیوتیک"},
    {"fa": "تتراسایکلین", "en": "Tetracycline", "type": "آنتی‌بیوتیک"},
    {"fa": "سفالکسین", "en": "Cefalexin", "type": "آنتی‌بیوتیک"},
    {"fa": "سیپروفلوکساسین", "en": "Ciprofloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "اریترومایسین", "en": "Erythromycin", "type": "آنتی‌بیوتیک"},
    {"fa": "تری‌متوپریم سولفا", "en": "Trimethoprim-Sulfa", "type": "آنتی‌بیوتیک"},
    {"fa": "فلورفنیکل", "en": "Florfenicol", "type": "آنتی‌بیوتیک"},
    {"fa": "جنتامایسین", "en": "Gentamicin", "type": "آنتی‌بیوتیک"},
    {"fa": "نئومایسین", "en": "Neomycin", "type": "آنتی‌بیوتیک"},
    {"fa": "کلرامفنیکل", "en": "Chloramphenicol", "type": "آنتی‌بیوتیک"},
    {"fa": "آمپی‌سیلین", "en": "Ampicillin", "type": "آنتی‌بیوتیک"},
    {"fa": "پنی‌سیلین جی", "en": "Penicillin G", "type": "آنتی‌بیوتیک"},
    {"fa": "پنی‌سیلین وی", "en": "Penicillin V", "type": "آنتی‌بیوتیک"},
    {"fa": "کلاوولانیک اسید", "en": "Clavulanic Acid", "type": "آنتی‌بیوتیک"},
    {"fa": "آموکسی‌کلاو", "en": "Amoxicillin-Clavulanate", "type": "آنتی‌بیوتیک"},
    {"fa": "سفتیوفور", "en": "Ceftiofur", "type": "آنتی‌بیوتیک"},
    {"fa": "سفازولین", "en": "Cefazolin", "type": "آنتی‌بیوتیک"},
    {"fa": "سفتازیدیم", "en": "Ceftazidime", "type": "آنتی‌بیوتیک"},
    {"fa": "سفوتاکسیم", "en": "Cefotaxime", "type": "آنتی‌بیوتیک"},
    {"fa": "سفتریاکسون", "en": "Ceftriaxone", "type": "آنتی‌بیوتیک"},
    {"fa": "سفپیم", "en": "Cefepime", "type": "آنتی‌بیوتیک"},
    {"fa": "سفاپیرین", "en": "Cefapirin", "type": "آنتی‌بیوتیک"},
    {"fa": "سفادروکسیل", "en": "Cefadroxil", "type": "آنتی‌بیوتیک"},
    {"fa": "ایمی‌پنم", "en": "Imipenem", "type": "آنتی‌بیوتیک"},
    {"fa": "مروپنم", "en": "Meropenem", "type": "آنتی‌بیوتیک"},
    {"fa": "آزیترومایسین", "en": "Azithromycin", "type": "آنتی‌بیوتیک"},
    {"fa": "کلاریترومایسین", "en": "Clarithromycin", "type": "آنتی‌بیوتیک"},
    {"fa": "تیلمیکوزین", "en": "Tilmicosin", "type": "آنتی‌بیوتیک"},
    {"fa": "توبرامایسین", "en": "Tobramycin", "type": "آنتی‌بیوتیک"},
    {"fa": "آمیکاسین", "en": "Amikacin", "type": "آنتی‌بیوتیک"},
    {"fa": "استرپتومایسین", "en": "Streptomycin", "type": "آنتی‌بیوتیک"},
    {"fa": "باسیتراسین", "en": "Bacitracin", "type": "آنتی‌بیوتیک"},
    {"fa": "پلی‌میکسین بی", "en": "Polymyxin B", "type": "آنتی‌بیوتیک"},
    {"fa": "کلیستین", "en": "Colistin", "type": "آنتی‌بیوتیک"},
    {"fa": "ریفامپین", "en": "Rifampin", "type": "آنتی‌بیوتیک"},
    {"fa": "وانکومایسین", "en": "Vancomycin", "type": "آنتی‌بیوتیک"},
    {"fa": "لینکومایسین", "en": "Lincomycin", "type": "آنتی‌بیوتیک"},
    {"fa": "کلیندامایسین", "en": "Clindamycin", "type": "آنتی‌بیوتیک"},
    {"fa": "اسپکتینومایسین", "en": "Spectinomycin", "type": "آنتی‌بیوتیک"},
    {"fa": "نیتروفورانتوئین", "en": "Nitrofurantoin", "type": "آنتی‌بیوتیک"},
    {"fa": "سولفادیازین", "en": "Sulfadiazine", "type": "آنتی‌بیوتیک"},
    {"fa": "سولفامتوکسازول", "en": "Sulfamethoxazole", "type": "آنتی‌بیوتیک"},
    {"fa": "سولفادیمتوکسین", "en": "Sulfadimethoxine", "type": "آنتی‌بیوتیک"},
    {"fa": "دانوفلوکساسین", "en": "Danofloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "مارب وفلوکساسین", "en": "Marbofloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "اورب وفلوکساسین", "en": "Orbifloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "پرادوفلوکساسین", "en": "Pradofloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "دی‌فلوکساسین", "en": "Difloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "نورفلوکساسین", "en": "Norfloxacin", "type": "آنتی‌بیوتیک"},
    {"fa": "اکسی‌تتراسایکلین", "en": "Oxytetracycline", "type": "آنتی‌بیوتیک"},
    {"fa": "کلرتتراسایکلین", "en": "Chlortetracycline", "type": "آنتی‌بیوتیک"},
    {"fa": "مینوسایکلین", "en": "Minocycline", "type": "آنتی‌بیوتیک"},
    {"fa": "تیامولین", "en": "Tiamulin", "type": "آنتی‌بیوتیک"},
    {"fa": "والنمولین", "en": "Valnemulin", "type": "آنتی‌بیوتیک"},
    {"fa": "انرامایسین", "en": "Enramycin", "type": "آنتی‌بیوتیک"},
    {"fa": "آویلامایسین", "en": "Avilamycin", "type": "آنتی‌بیوتیک"},
    {"fa": "بامبرمایسین", "en": "Bambermycin", "type": "آنتی‌بیوتیک"},
    {"fa": "مترونیدازول", "en": "Metronidazole", "type": "ضد انگل"},
    {"fa": "فنبندازول", "en": "Fenbendazole", "type": "ضد انگل"},
    {"fa": "آلبندازول", "en": "Albendazole", "type": "ضد انگل"},
    {"fa": "لوامیزول", "en": "Levamisole", "type": "ضد انگل"},
    {"fa": "ایورمکتین", "en": "Ivermectin", "type": "ضد انگل"},
    {"fa": "پرازیکوانتل", "en": "Praziquantel", "type": "ضد انگل"},
    {"fa": "تولترازوریل", "en": "Toltrazuril", "type": "ضد انگل"},
    {"fa": "آمپرولیوم", "en": "Amprolium", "type": "ضد انگل"},
    {"fa": "مبندازول", "en": "Mebendazole", "type": "ضد انگل"},
    {"fa": "تیابندازول", "en": "Thiabendazole", "type": "ضد انگل"},
    {"fa": "اکسفندازول", "en": "Oxfendazole", "type": "ضد انگل"},
    {"fa": "فبانتل", "en": "Febantel", "type": "ضد انگل"},
    {"fa": "پیرانتل", "en": "Pyrantel", "type": "ضد انگل"},
    {"fa": "مورانتل", "en": "Morantel", "type": "ضد انگل"},
    {"fa": "دورامکتین", "en": "Doramectin", "type": "ضد انگل"},
    {"fa": "اپرینومکتین", "en": "Eprinomectin", "type": "ضد انگل"},
    {"fa": "مکسیدکتین", "en": "Moxidectin", "type": "ضد انگل"},
    {"fa": "سلامکتین", "en": "Selamectin", "type": "ضد انگل"},
    {"fa": "میلبمایسین", "en": "Milbemycin", "type": "ضد انگل"},
    {"fa": "دکوکینات", "en": "Decoquinate", "type": "ضد انگل"},
    {"fa": "دیکلازوریل", "en": "Diclazuril", "type": "ضد انگل"},
    {"fa": "مادورامایسین", "en": "Maduramicin", "type": "ضد انگل"},
    {"fa": "موننزین", "en": "Monensin", "type": "ضد انگل"},
    {"fa": "سالینومایسین", "en": "Salinomycin", "type": "ضد انگل"},
    {"fa": "نیکاربازین", "en": "Nicarbazin", "type": "ضد انگل"},
    {"fa": "لازالوسید", "en": "Lasalocid", "type": "ضد انگل"},
    {"fa": "رابنیدین", "en": "Robenidine", "type": "ضد انگل"},
    {"fa": "تینیدازول", "en": "Tinidazole", "type": "ضد انگل"},
    {"fa": "رونیدازول", "en": "Ronidazole", "type": "ضد انگل"},
    {"fa": "دیمتریدازول", "en": "Dimetridazole", "type": "ضد انگل"},
    {"fa": "کلوپیدول", "en": "Clopidol", "type": "ضد انگل"},
    {"fa": "فیپرونیل", "en": "Fipronil", "type": "ضد انگل"},
    {"fa": "ایمیداکلوپرید", "en": "Imidacloprid", "type": "ضد انگل"},
    {"fa": "پرمترین", "en": "Permethrin", "type": "ضد انگل"},
    {"fa": "آمیتراز", "en": "Amitraz", "type": "ضد انگل"},
    {"fa": "کاربریل", "en": "Carbaryl", "type": "ضد انگل"},
    {"fa": "لوفنورون", "en": "Lufenuron", "type": "ضد انگل"},
    {"fa": "نیتنپیرام", "en": "Nitenpyram", "type": "ضد انگل"},
    {"fa": "اسپینوساد", "en": "Spinosad", "type": "ضد انگل"},
    {"fa": "فلورالانر", "en": "Fluralaner", "type": "ضد انگل"},
    {"fa": "آفوکسولانر", "en": "Afoxolaner", "type": "ضد انگل"},
    {"fa": "ساروالنر", "en": "Sarolaner", "type": "ضد انگل"},
    {"fa": "لوتیلانر", "en": "Lotilaner", "type": "ضد انگل"},
    {"fa": "ملوکسیکام", "en": "Meloxicam", "type": "ضد التهاب"},
    {"fa": "کتوپروفن", "en": "Ketoprofen", "type": "ضد التهاب"},
    {"fa": "فلونیکسین مگلومین", "en": "Flunixin Meglumine", "type": "ضد التهاب"},
    {"fa": "کارپروفن", "en": "Carprofen", "type": "ضد التهاب"},
    {"fa": "فنیل‌بوتازون", "en": "Phenylbutazone", "type": "ضد التهاب"},
    {"fa": "آسپرین", "en": "Aspirin", "type": "ضد التهاب"},
    {"fa": "دراکوکسیب", "en": "Deracoxib", "type": "ضد التهاب"},
    {"fa": "فیروکوکسیب", "en": "Firocoxib", "type": "ضد التهاب"},
    {"fa": "روبناکوکسیب", "en": "Robenacoxib", "type": "ضد التهاب"},
    {"fa": "مکلوفنامیک اسید", "en": "Meclofenamic Acid", "type": "ضد التهاب"},
    {"fa": "پیروکسیکام", "en": "Piroxicam", "type": "ضد التهاب"},
    {"fa": "تپوکسالین", "en": "Tepoxalin", "type": "ضد التهاب"},
    {"fa": "اتودولاک", "en": "Etodolac", "type": "ضد التهاب"},
    {"fa": "ناپروکسن", "en": "Naproxen", "type": "ضد التهاب"},
    {"fa": "ایبوپروفن", "en": "Ibuprofen", "type": "ضد التهاب"},
    {"fa": "دیمتیل سولفوکسید", "en": "Dimethyl Sulfoxide", "type": "ضد التهاب"},
    {"fa": "دگزامتازون", "en": "Dexamethasone", "type": "کورتیکواستروئید"},
    {"fa": "پردنیزولون", "en": "Prednisolone", "type": "کورتیکواستروئید"},
    {"fa": "هیدروکورتیزون", "en": "Hydrocortisone", "type": "کورتیکواستروئید"},
    {"fa": "پردنیزون", "en": "Prednisone", "type": "کورتیکواستروئید"},
    {"fa": "تریامسینولون", "en": "Triamcinolone", "type": "کورتیکواستروئید"},
    {"fa": "متیل‌پردنیزولون", "en": "Methylprednisolone", "type": "کورتیکواستروئید"},
    {"fa": "بتامتازون", "en": "Betamethasone", "type": "کورتیکواستروئید"},
    {"fa": "فلومتازون", "en": "Flumethasone", "type": "کورتیکواستروئید"},
    {"fa": "ایزوفلوپردون", "en": "Isoflupredone", "type": "کورتیکواستروئید"},
    {"fa": "بودزوناید", "en": "Budesonide", "type": "کورتیکواستروئید"},
    {"fa": "فلوتیکازون", "en": "Fluticasone", "type": "کورتیکواستروئید"},
    {"fa": "کلرفنیرامین", "en": "Chlorpheniramine", "type": "ضد حساسیت"},
    {"fa": "دیفن‌هیدرامین", "en": "Diphenhydramine", "type": "ضد حساسیت"},
    {"fa": "هیدروکسی‌زین", "en": "Hydroxyzine", "type": "ضد حساسیت"},
    {"fa": "ستیریزین", "en": "Cetirizine", "type": "ضد حساسیت"},
    {"fa": "لوراتادین", "en": "Loratadine", "type": "ضد حساسیت"},
    {"fa": "فکسوفنادین", "en": "Fexofenadine", "type": "ضد حساسیت"},
    {"fa": "سیپروهپتادین", "en": "Cyproheptadine", "type": "ضد حساسیت"},
    {"fa": "کلماستین", "en": "Clemastine", "type": "ضد حساسیت"},
    {"fa": "ماروپیتانت", "en": "Maropitant", "type": "ضد حساسیت"},
    {"fa": "کتامین", "en": "Ketamine", "type": "بیهوشی"},
    {"fa": "زایلازین", "en": "Xylazine", "type": "بیهوشی"},
    {"fa": "ایزوفلوران", "en": "Isoflurane", "type": "بیهوشی"},
    {"fa": "دیازپام", "en": "Diazepam", "type": "بیهوشی"},
    {"fa": "میدازولام", "en": "Midazolam", "type": "بیهوشی"},
    {"fa": "پروپوفول", "en": "Propofol", "type": "بیهوشی"},
    {"fa": "تیلتامین-زولازپام", "en": "Tiletamine-Zolazepam", "type": "بیهوشی"},
    {"fa": "آلفاکسالون", "en": "Alfaxalone", "type": "بیهوشی"},
    {"fa": "سووفلوران", "en": "Sevoflurane", "type": "بیهوشی"},
    {"fa": "تیوپنتال", "en": "Thiopental", "type": "بیهوشی"},
    {"fa": "پنتوباربیتال", "en": "Pentobarbital", "type": "بیهوشی"},
    {"fa": "مدتومیدین", "en": "Medetomidine", "type": "بیهوشی"},
    {"fa": "دکسمدتومیدین", "en": "Dexmedetomidine", "type": "بیهوشی"},
    {"fa": "آسپرومازین", "en": "Acepromazine", "type": "بیهوشی"},
    {"fa": "آتیپامزول", "en": "Atipamezole", "type": "بیهوشی"},
    {"fa": "فلومازنیل", "en": "Flumazenil", "type": "بیهوشی"},
    {"fa": "دوکساپرام", "en": "Doxapram", "type": "بیهوشی"},
    {"fa": "یوهمبین", "en": "Yohimbine", "type": "بیهوشی"},
    {"fa": "گلیکوپیرولات", "en": "Glycopyrrolate", "type": "بیهوشی"},
    {"fa": "فنوباربیتال", "en": "Phenobarbital", "type": "بیهوشی"},
    {"fa": "کلرال هیدرات", "en": "Chloral Hydrate", "type": "بیهوشی"},
    {"fa": "رومیفیدین", "en": "Romifidine", "type": "بیهوشی"},
    {"fa": "دتومیدین", "en": "Detomidine", "type": "بیهوشی"},
    {"fa": "گوایفنزین", "en": "Guaifenesin", "type": "بیهوشی"},
    {"fa": "بوتورفانول", "en": "Butorphanol", "type": "ضد درد"},
    {"fa": "ترامادول", "en": "Tramadol", "type": "ضد درد"},
    {"fa": "گاباپنتین", "en": "Gabapentin", "type": "ضد درد"},
    {"fa": "مورفین", "en": "Morphine", "type": "ضد درد"},
    {"fa": "فنتانیل", "en": "Fentanyl", "type": "ضد درد"},
    {"fa": "بوپرنورفین", "en": "Buprenorphine", "type": "ضد درد"},
    {"fa": "هیدرومورفون", "en": "Hydromorphone", "type": "ضد درد"},
    {"fa": "متادون", "en": "Methadone", "type": "ضد درد"},
    {"fa": "نالوکسان", "en": "Naloxone", "type": "ضد درد"},
    {"fa": "نالترکسون", "en": "Naltrexone", "type": "ضد درد"},
    {"fa": "لیدوکائین", "en": "Lidocaine", "type": "ضد درد"},
    {"fa": "بوپیواکائین", "en": "Bupivacaine", "type": "ضد درد"},
    {"fa": "مپیواکائین", "en": "Mepivacaine", "type": "ضد درد"},
    {"fa": "پتیدین", "en": "Meperidine", "type": "ضد درد"},
    {"fa": "آمانتادین", "en": "Amantadine", "type": "ضد درد"},
    {"fa": "پرگابالین", "en": "Pregabalin", "type": "ضد درد"},
    {"fa": "فلوکونازول", "en": "Fluconazole", "type": "ضد قارچ"},
    {"fa": "نیستاتین", "en": "Nystatin", "type": "ضد قارچ"},
    {"fa": "کتوکونازول", "en": "Ketoconazole", "type": "ضد قارچ"},
    {"fa": "آمفوتریسین بی", "en": "Amphotericin B", "type": "ضد قارچ"},
    {"fa": "ایتراکونازول", "en": "Itraconazole", "type": "ضد قارچ"},
    {"fa": "وریکونازول", "en": "Voriconazole", "type": "ضد قارچ"},
    {"fa": "پوزاکونازول", "en": "Posaconazole", "type": "ضد قارچ"},
    {"fa": "تربینافین", "en": "Terbinafine", "type": "ضد قارچ"},
    {"fa": "گریزئوفولوین", "en": "Griseofulvin", "type": "ضد قارچ"},
    {"fa": "کلوتریمازول", "en": "Clotrimazole", "type": "ضد قارچ"},
    {"fa": "میکونازول", "en": "Miconazole", "type": "ضد قارچ"},
    {"fa": "کاسپوفانژین", "en": "Caspofungin", "type": "ضد قارچ"},
    {"fa": "فلوسیتوزین", "en": "Flucytosine", "type": "ضد قارچ"},
    {"fa": "مولتی‌ویتامین", "en": "Multivitamin", "type": "تقویتی"},
    {"fa": "ویتامین A", "en": "Vitamin A", "type": "تقویتی"},
    {"fa": "ویتامین B1", "en": "Thiamine", "type": "تقویتی"},
    {"fa": "ویتامین B2", "en": "Riboflavin", "type": "تقویتی"},
    {"fa": "ویتامین B3", "en": "Niacin", "type": "تقویتی"},
    {"fa": "ویتامین B5", "en": "Pantothenic Acid", "type": "تقویتی"},
    {"fa": "ویتامین B6", "en": "Pyridoxine", "type": "تقویتی"},
    {"fa": "ویتامین B9", "en": "Folic Acid", "type": "تقویتی"},
    {"fa": "ویتامین B12", "en": "Cyanocobalamin", "type": "تقویتی"},
    {"fa": "ویتامین B کمپلکس", "en": "Vitamin B Complex", "type": "تقویتی"},
    {"fa": "ویتامین C", "en": "Vitamin C", "type": "تقویتی"},
    {"fa": "ویتامین D3", "en": "Vitamin D3", "type": "تقویتی"},
    {"fa": "ویتامین E", "en": "Vitamin E", "type": "تقویتی"},
    {"fa": "ویتامین E سلنیوم", "en": "Vitamin E + Selenium", "type": "تقویتی"},
    {"fa": "ویتامین K1", "en": "Phytonadione", "type": "تقویتی"},
    {"fa": "ویتامین K3", "en": "Menadione", "type": "تقویتی"},
    {"fa": "کلسیم", "en": "Calcium", "type": "تقویتی"},
    {"fa": "کلسیم گلوکونات", "en": "Calcium Gluconate", "type": "تقویتی"},
    {"fa": "کلسیم بوروگلوکونات", "en": "Calcium Borogluconate", "type": "تقویتی"},
    {"fa": "فسفر", "en": "Phosphorus", "type": "تقویتی"},
    {"fa": "آهن دکستران", "en": "Iron Dextran", "type": "تقویتی"},
    {"fa": "روی", "en": "Zinc", "type": "تقویتی"},
    {"fa": "منگنز", "en": "Manganese", "type": "تقویتی"},
    {"fa": "مس", "en": "Copper", "type": "تقویتی"},
    {"fa": "سلنیوم", "en": "Selenium", "type": "تقویتی"},
    {"fa": "ید", "en": "Iodine", "type": "تقویتی"},
    {"fa": "الکترولیت", "en": "Electrolyte", "type": "تقویتی"},
    {"fa": "پروبیوتیک", "en": "Probiotic", "type": "تقویتی"},
    {"fa": "پری‌بیوتیک", "en": "Prebiotic", "type": "تقویتی"},
    {"fa": "اسید آمینه", "en": "Amino Acid", "type": "تقویتی"},
    {"fa": "متیونین", "en": "Methionine", "type": "تقویتی"},
    {"fa": "لیزین", "en": "Lysine", "type": "تقویتی"},
    {"fa": "کولین", "en": "Choline", "type": "تقویتی"},
    {"fa": "بیوتین", "en": "Biotin", "type": "تقویتی"},
    {"fa": "اسید لینولئیک", "en": "Linoleic Acid", "type": "تقویتی"},
    {"fa": "امگا ۳", "en": "Omega-3", "type": "تقویتی"},
    {"fa": "روغن ماهی", "en": "Fish Oil", "type": "تقویتی"},
    {"fa": "گلوکزامین", "en": "Glucosamine", "type": "تقویتی"},
    {"fa": "کندرویتین", "en": "Chondroitin", "type": "تقویتی"},
    {"fa": "ال-کارنیتین", "en": "L-Carnitine", "type": "تقویتی"},
    {"fa": "تائورین", "en": "Taurine", "type": "تقویتی"},
    {"fa": "کو-آنزیم Q10", "en": "Coenzyme Q10", "type": "تقویتی"},
    {"fa": "سام-ایی", "en": "SAMe", "type": "تقویتی"},
    {"fa": "سیلی‌مارین", "en": "Silymarin", "type": "تقویتی"},
    {"fa": "سرم نمکی", "en": "Normal Saline", "type": "سرم و مایعات"},
    {"fa": "سرم قندی", "en": "Dextrose Saline", "type": "سرم و مایعات"},
    {"fa": "رینگر لاکتات", "en": "Ringer Lactate", "type": "سرم و مایعات"},
    {"fa": "دکستروز ۵٪", "en": "Dextrose 5%", "type": "سرم و مایعات"},
    {"fa": "دکستروز ۵۰٪", "en": "Dextrose 50%", "type": "سرم و مایعات"},
    {"fa": "هتااستارچ", "en": "Hetastarch", "type": "سرم و مایعات"},
    {"fa": "مانیتول", "en": "Mannitol", "type": "سرم و مایعات"},
    {"fa": "سدیم بیکربنات", "en": "Sodium Bicarbonate", "type": "سرم و مایعات"},
    {"fa": "پتاسیم کلراید", "en": "Potassium Chloride", "type": "سرم و مایعات"},
    {"fa": "فوروزماید", "en": "Furosemide", "type": "قلبی‌عروقی"},
    {"fa": "دیگوکسین", "en": "Digoxin", "type": "قلبی‌عروقی"},
    {"fa": "اناالپریل", "en": "Enalapril", "type": "قلبی‌عروقی"},
    {"fa": "بنازپریل", "en": "Benazepril", "type": "قلبی‌عروقی"},
    {"fa": "پیموبندان", "en": "Pimobendan", "type": "قلبی‌عروقی"},
    {"fa": "آملودیپین", "en": "Amlodipine", "type": "قلبی‌عروقی"},
    {"fa": "آتنولول", "en": "Atenolol", "type": "قلبی‌عروقی"},
    {"fa": "پروپرانولول", "en": "Propranolol", "type": "قلبی‌عروقی"},
    {"fa": "دیلتیازم", "en": "Diltiazem", "type": "قلبی‌عروقی"},
    {"fa": "هیدرالازین", "en": "Hydralazine", "type": "قلبی‌عروقی"},
    {"fa": "نیتروگلیسیرین", "en": "Nitroglycerin", "type": "قلبی‌عروقی"},
    {"fa": "نیتروپروساید", "en": "Nitroprusside", "type": "قلبی‌عروقی"},
    {"fa": "دوبوتامین", "en": "Dobutamine", "type": "قلبی‌عروقی"},
    {"fa": "دوپامین", "en": "Dopamine", "type": "قلبی‌عروقی"},
    {"fa": "اسپیرونولاکتون", "en": "Spironolactone", "type": "قلبی‌عروقی"},
    {"fa": "هیدروکلروتیازید", "en": "Hydrochlorothiazide", "type": "قلبی‌عروقی"},
    {"fa": "تورسماید", "en": "Torsemide", "type": "قلبی‌عروقی"},
    {"fa": "وراپامیل", "en": "Verapamil", "type": "قلبی‌عروقی"},
    {"fa": "سوتالول", "en": "Sotalol", "type": "قلبی‌عروقی"},
    {"fa": "پروکائین‌آمید", "en": "Procainamide", "type": "قلبی‌عروقی"},
    {"fa": "لیزینوپریل", "en": "Lisinopril", "type": "قلبی‌عروقی"},
    {"fa": "تلمیسارتان", "en": "Telmisartan", "type": "قلبی‌عروقی"},
    {"fa": "سیلدنافیل", "en": "Sildenafil", "type": "قلبی‌عروقی"},
    {"fa": "متوکلوپرامید", "en": "Metoclopramide", "type": "گوارشی"},
    {"fa": "رانیتیدین", "en": "Ranitidine", "type": "گوارشی"},
    {"fa": "امپرازول", "en": "Omeprazole", "type": "گوارشی"},
    {"fa": "فاموتیدین", "en": "Famotidine", "type": "گوارشی"},
    {"fa": "سوکرالفات", "en": "Sucralfate", "type": "گوارشی"},
    {"fa": "میزوپروستول", "en": "Misoprostol", "type": "گوارشی"},
    {"fa": "سیزاپراید", "en": "Cisapride", "type": "گوارشی"},
    {"fa": "اوندانسترون", "en": "Ondansetron", "type": "گوارشی"},
    {"fa": "ماروپیتانت سیترات", "en": "Maropitant Citrate", "type": "گوارشی"},
    {"fa": "دولاسترون", "en": "Dolasetron", "type": "گوارشی"},
    {"fa": "متوکلوپرامید", "en": "Metoclopramide", "type": "گوارشی"},
    {"fa": "لوپرامید", "en": "Loperamide", "type": "گوارشی"},
    {"fa": "بیسموت ساب‌سالیسیلات", "en": "Bismuth Subsalicylate", "type": "گوارشی"},
    {"fa": "لاکتولوز", "en": "Lactulose", "type": "گوارشی"},
    {"fa": "پسیلیوم", "en": "Psyllium", "type": "گوارشی"},
    {"fa": "داکوزات سدیم", "en": "Docusate Sodium", "type": "گوارشی"},
    {"fa": "پانکراتین", "en": "Pancreatin", "type": "گوارشی"},
    {"fa": "پانکرلیپاز", "en": "Pancrelipase", "type": "گوارشی"},
    {"fa": "اسلانزاپین", "en": "Esomeprazole", "type": "گوارشی"},
    {"fa": "لانزوپرازول", "en": "Lansoprazole", "type": "گوارشی"},
    {"fa": "سایمتیکون", "en": "Simethicone", "type": "گوارشی"},
    {"fa": "شارکل فعال", "en": "Activated Charcoal", "type": "ضد سم"},
    {"fa": "آتروپین", "en": "Atropine", "type": "آنتی‌دوت"},
    {"fa": "پرالیدوکسیم", "en": "Pralidoxime", "type": "آنتی‌دوت"},
    {"fa": "دیمرکاپرول", "en": "Dimercaprol", "type": "آنتی‌دوت"},
    {"fa": "اتیلن دیامین تترا استیک اسید", "en": "EDTA", "type": "آنتی‌دوت"},
    {"fa": "دفروکسامین", "en": "Deferoxamine", "type": "آنتی‌دوت"},
    {"fa": "ان-استیل سیستئین", "en": "N-Acetylcysteine", "type": "آنتی‌دوت"},
    {"fa": "اتانول", "en": "Ethanol", "type": "آنتی‌دوت"},
    {"fa": "فومپیزول", "en": "Fomepizole", "type": "آنتی‌دوت"},
    {"fa": "اپی‌نفرین", "en": "Epinephrine", "type": "اورژانسی"},
    {"fa": "وازوپرسین", "en": "Vasopressin", "type": "اورژانسی"},
    {"fa": "آمینوفیلین", "en": "Aminophylline", "type": "اورژانسی"},
    {"fa": "تئوفیلین", "en": "Theophylline", "type": "اورژانسی"},
    {"fa": "ترپنتالین", "en": "Terbutaline", "type": "اورژانسی"},
    {"fa": "آلبوترول", "en": "Albuterol", "type": "اورژانسی"},
    {"fa": "اکسی‌توسین", "en": "Oxytocin", "type": "هورمونی"},
    {"fa": "دسموپرسین", "en": "Desmopressin", "type": "هورمونی"},
    {"fa": "انسولین", "en": "Insulin", "type": "هورمونی"},
    {"fa": "لووتیروکسین", "en": "Levothyroxine", "type": "هورمونی"},
    {"fa": "متی‌مازول", "en": "Methimazole", "type": "هورمونی"},
    {"fa": "تستوسترون", "en": "Testosterone", "type": "هورمونی"},
    {"fa": "استرادیول", "en": "Estradiol", "type": "هورمونی"},
    {"fa": "پروژسترون", "en": "Progesterone", "type": "هورمونی"},
    {"fa": "مژسترول", "en": "Megestrol", "type": "هورمونی"},
    {"fa": "آلترنوژست", "en": "Altrenogest", "type": "هورمونی"},
    {"fa": "GnRH آنالوگ", "en": "GnRH Analog", "type": "هورمونی"},
    {"fa": "HCG", "en": "HCG", "type": "هورمونی"},
    {"fa": "پروستاگلاندین F2α", "en": "Prostaglandin F2a", "type": "هورمونی"},
    {"fa": "کلوپروستنول", "en": "Cloprostenol", "type": "هورمونی"},
    {"fa": "دینوپروست", "en": "Dinoprost", "type": "هورمونی"},
    {"fa": "تریلوستان", "en": "Trilostane", "type": "هورمونی"},
    {"fa": "میتوتان", "en": "Mitotane", "type": "هورمونی"},
    {"fa": "دسلورلین", "en": "Deslorelin", "type": "هورمونی"},
    {"fa": "کابرگولین", "en": "Cabergoline", "type": "هورمونی"},
    {"fa": "بروموکریپتین", "en": "Bromocriptine", "type": "هورمونی"},
    {"fa": "سوماتوتروپین", "en": "Somatotropin", "type": "هورمونی"},
    {"fa": "کالسیتونین", "en": "Calcitonin", "type": "هورمونی"},
    {"fa": "فنوباربیتال", "en": "Phenobarbital", "type": "ضد تشنج"},
    {"fa": "پتاسیم بروماید", "en": "Potassium Bromide", "type": "ضد تشنج"},
    {"fa": "لوتیراستام", "en": "Levetiracetam", "type": "ضد تشنج"},
    {"fa": "زونیساماید", "en": "Zonisamide", "type": "ضد تشنج"},
    {"fa": "کلونازپام", "en": "Clonazepam", "type": "ضد تشنج"},
    {"fa": "فلبامات", "en": "Felbamate", "type": "ضد تشنج"},
    {"fa": "توپیرامات", "en": "Topiramate", "type": "ضد تشنج"},
    {"fa": "ایمپیتوئین", "en": "Imepitoin", "type": "ضد تشنج"},
    {"fa": "فلوکستین", "en": "Fluoxetine", "type": "رفتاری"},
    {"fa": "کلومیپرامین", "en": "Clomipramine", "type": "رفتاری"},
    {"fa": "آمیتریپتیلین", "en": "Amitriptyline", "type": "رفتاری"},
    {"fa": "ترازودون", "en": "Trazodone", "type": "رفتاری"},
    {"fa": "سرترالین", "en": "Sertraline", "type": "رفتاری"},
    {"fa": "پاروکستین", "en": "Paroxetine", "type": "رفتاری"},
    {"fa": "بوسپیرون", "en": "Buspirone", "type": "رفتاری"},
    {"fa": "آلپرازولام", "en": "Alprazolam", "type": "رفتاری"},
    {"fa": "سلژیلین", "en": "Selegiline", "type": "رفتاری"},
    {"fa": "دوکسپین", "en": "Doxepin", "type": "رفتاری"},
    {"fa": "سیکلوسپورین", "en": "Cyclosporine", "type": "ایمنی"},
    {"fa": "آزاتیوپرین", "en": "Azathioprine", "type": "ایمنی"},
    {"fa": "مایکوفنولات", "en": "Mycophenolate", "type": "ایمنی"},
    {"fa": "سیکلوفسفامید", "en": "Cyclophosphamide", "type": "ضد سرطان"},
    {"fa": "وین‌کریستین", "en": "Vincristine", "type": "ضد سرطان"},
    {"fa": "داکسوروبیسین", "en": "Doxorubicin", "type": "ضد سرطان"},
    {"fa": "کاربوپلاتین", "en": "Carboplatin", "type": "ضد سرطان"},
    {"fa": "سیس‌پلاتین", "en": "Cisplatin", "type": "ضد سرطان"},
    {"fa": "کلرامبوسیل", "en": "Chlorambucil", "type": "ضد سرطان"},
    {"fa": "لوموستین", "en": "Lomustine", "type": "ضد سرطان"},
    {"fa": "ملفالان", "en": "Melphalan", "type": "ضد سرطان"},
    {"fa": "متوترکسات", "en": "Methotrexate", "type": "ضد سرطان"},
    {"fa": "توسرانیب", "en": "Toceranib", "type": "ضد سرطان"},
    {"fa": "ماسیتینیب", "en": "Masitinib", "type": "ضد سرطان"},
    {"fa": "تیمولول", "en": "Timolol", "type": "چشمی"},
    {"fa": "دورزولامید", "en": "Dorzolamide", "type": "چشمی"},
    {"fa": "لاتانوپروست", "en": "Latanoprost", "type": "چشمی"},
    {"fa": "تروپیکامید", "en": "Tropicamide", "type": "چشمی"},
    {"fa": "فنیل‌افرین چشمی", "en": "Phenylephrine Ophthalmic", "type": "چشمی"},
    {"fa": "سیکلوپنتولات", "en": "Cyclopentolate", "type": "چشمی"},
    {"fa": "فلوربیپروفن چشمی", "en": "Flurbiprofen Ophthalmic", "type": "چشمی"},
    {"fa": "دیکلوفناک چشمی", "en": "Diclofenac Ophthalmic", "type": "چشمی"},
    {"fa": "سیپروفلوکساسین چشمی", "en": "Ciprofloxacin Ophthalmic", "type": "چشمی"},
    {"fa": "جنتامایسین چشمی", "en": "Gentamicin Ophthalmic", "type": "چشمی"},
    {"fa": "اکسی‌فلوکساسین", "en": "Ofloxacin", "type": "چشمی"},
    {"fa": "تری‌پل آنتی‌بیوتیک چشمی", "en": "Triple Antibiotic Ophthalmic", "type": "چشمی"},
    {"fa": "کلرهگزیدین", "en": "Chlorhexidine", "type": "پوستی"},
    {"fa": "پوویدون ید", "en": "Povidone Iodine", "type": "پوستی"},
    {"fa": "سولفور", "en": "Sulfur", "type": "پوستی"},
    {"fa": "بنزوئیل پراکسید", "en": "Benzoyl Peroxide", "type": "پوستی"},
    {"fa": "موپیروسین", "en": "Mupirocin", "type": "پوستی"},
    {"fa": "سیلور سولفادیازین", "en": "Silver Sulfadiazine", "type": "پوستی"},
    {"fa": "تاکرولیموس", "en": "Tacrolimus", "type": "پوستی"},
    {"fa": "هیدروژن پراکسید", "en": "Hydrogen Peroxide", "type": "پوستی"},
    {"fa": "برمهگزین", "en": "Bromhexine", "type": "تنفسی"},
    {"fa": "آمبروکسول", "en": "Ambroxol", "type": "تنفسی"},
    {"fa": "دکسترومتورفان", "en": "Dextromethorphan", "type": "تنفسی"},
    {"fa": "هیدروکدون", "en": "Hydrocodone", "type": "تنفسی"},
    {"fa": "کدئین", "en": "Codeine", "type": "تنفسی"},
    {"fa": "آستیل سیستئین", "en": "Acetylcysteine", "type": "تنفسی"},
    {"fa": "فنازوپیریدین", "en": "Phenazopyridine", "type": "ادراری"},
    {"fa": "فنوکسی‌بنزامین", "en": "Phenoxybenzamine", "type": "ادراری"},
    {"fa": "پرازوسین", "en": "Prazosin", "type": "ادراری"},
    {"fa": "تامسولوسین", "en": "Tamsulosin", "type": "ادراری"},
    {"fa": "بتانکول", "en": "Bethanechol", "type": "ادراری"},
    {"fa": "دیات‌آمینوپیریمیدین", "en": "Diethylstilbestrol", "type": "ادراری"},
    {"fa": "فنیل‌پروپانولامین", "en": "Phenylpropanolamine", "type": "ادراری"},
    {"fa": "آلوپورینول", "en": "Allopurinol", "type": "متفرقه"},
    {"fa": "کلشی‌سین", "en": "Colchicine", "type": "متفرقه"},
    {"fa": "پنتوکسی‌فیلین", "en": "Pentoxifylline", "type": "متفرقه"},
    {"fa": "دانترولن", "en": "Dantrolene", "type": "متفرقه"},
    {"fa": "متوکاربامول", "en": "Methocarbamol", "type": "متفرقه"},
    {"fa": "باکلوفن", "en": "Baclofen", "type": "متفرقه"},
    {"fa": "اورسودیول", "en": "Ursodiol", "type": "متفرقه"},
    {"fa": "مزالازین", "en": "Mesalamine", "type": "متفرقه"},
    {"fa": "سولفاسالازین", "en": "Sulfasalazine", "type": "متفرقه"},
    {"fa": "کرومولین سدیم", "en": "Cromolyn Sodium", "type": "متفرقه"},
    {"fa": "ایپراتروپیوم", "en": "Ipratropium", "type": "متفرقه"},
    {"fa": "منتول", "en": "Menthol", "type": "متفرقه"},
    {"fa": "کافئین", "en": "Caffeine", "type": "متفرقه"},
    {"fa": "آمونیوم کلراید", "en": "Ammonium Chloride", "type": "متفرقه"},
    {"fa": "آمینوگلیکوزید ترکیبی", "en": "Aminoglycoside Compound", "type": "آنتی‌بیوتیک"},
    {"fa": "سفوپرازون", "en": "Cefoperazone", "type": "آنتی‌بیوتیک"},
    {"fa": "دورکسین", "en": "Doxyvet", "type": "آنتی‌بیوتیک"},
    {"fa": "تایلان", "en": "Tylan", "type": "آنتی‌بیوتیک"},
    {"fa": "آنتی‌بیوتیک تزریقی LA", "en": "Long-Acting Antibiotic", "type": "آنتی‌بیوتیک"},
    {"fa": "پنی‌سیلین-استرپتومایسین", "en": "Pen-Strep", "type": "آنتی‌بیوتیک"},
    {"fa": "فسفومایسین", "en": "Fosfomycin", "type": "آنتی‌بیوتیک"},
    {"fa": "نوبیوسین", "en": "Novobiocin", "type": "آنتی‌بیوتیک"},
    {"fa": "دیکرونازول", "en": "Dicronazole", "type": "ضد انگل"},
    {"fa": "تری‌کلابندازول", "en": "Triclabendazole", "type": "ضد انگل"},
    {"fa": "نیکلوزامید", "en": "Niclosamide", "type": "ضد انگل"},
    {"fa": "اکسی‌کلوزاناید", "en": "Oxyclozanide", "type": "ضد انگل"},
    {"fa": "رافوکسانید", "en": "Rafoxanide", "type": "ضد انگل"},
    {"fa": "کلوزانتل", "en": "Closantel", "type": "ضد انگل"},
    {"fa": "نافتالوفوس", "en": "Naftalofos", "type": "ضد انگل"},
    {"fa": "هالوفوژنون", "en": "Halofuginone", "type": "ضد انگل"},
    {"fa": "دکسپانتنول", "en": "Dexpanthenol", "type": "تقویتی"},
    {"fa": "بوتافسفان", "en": "Butafosfan", "type": "تقویتی"},
    {"fa": "کتوزاید", "en": "Ketoaid", "type": "تقویتی"},
    {"fa": "هماتینیک", "en": "Hematinic", "type": "تقویتی"},
    {"fa": "فروس سولفات", "en": "Ferrous Sulfate", "type": "تقویتی"},
    {"fa": "ویتامین AD3E", "en": "Vitamin AD3E", "type": "تقویتی"},
    {"fa": "ویتامین AD3EC", "en": "Vitamin AD3EC", "type": "تقویتی"},
    {"fa": "کبالت", "en": "Cobalt", "type": "تقویتی"},
    {"fa": "متوپرولول", "en": "Metoprolol", "type": "قلبی‌عروقی"},
    {"fa": "کارودیلول", "en": "Carvedilol", "type": "قلبی‌عروقی"},
    {"fa": "رامیپریل", "en": "Ramipril", "type": "قلبی‌عروقی"},
    {"fa": "کاپتوپریل", "en": "Captopril", "type": "قلبی‌عروقی"},
    {"fa": "پنتوپرازول", "en": "Pantoprazole", "type": "گوارشی"},
    {"fa": "دم‌پریدون", "en": "Domperidone", "type": "گوارشی"},
    {"fa": "اریتروپوئیتین", "en": "Erythropoietin", "type": "هورمونی"},
    {"fa": "تیروتروپین", "en": "Thyrotropin", "type": "هورمونی"},
    {"fa": "کورتیکوتروپین", "en": "Corticotropin", "type": "هورمونی"},
    {"fa": "فینسترید", "en": "Finasteride", "type": "هورمونی"},
    {"fa": "اتامسیلات", "en": "Etamsylate", "type": "اورژانسی"},
    {"fa": "ترانگزامیک اسید", "en": "Tranexamic Acid", "type": "اورژانسی"},
    {"fa": "آمینوکاپروئیک اسید", "en": "Aminocaproic Acid", "type": "اورژانسی"},
    {"fa": "هپارین", "en": "Heparin", "type": "قلبی‌عروقی"},
    {"fa": "وارفارین", "en": "Warfarin", "type": "قلبی‌عروقی"},
    {"fa": "کلوپیدوگرل", "en": "Clopidogrel", "type": "قلبی‌عروقی"},
    {"fa": "داناپاروئید", "en": "Danaparoid", "type": "قلبی‌عروقی"},
    {"fa": "آسیکلوویر", "en": "Acyclovir", "type": "ضد ویروس"},
    {"fa": "فامسیکلوویر", "en": "Famciclovir", "type": "ضد ویروس"},
    {"fa": "اینترفرون", "en": "Interferon", "type": "ضد ویروس"},
    {"fa": "اوسلتامیویر", "en": "Oseltamivir", "type": "ضد ویروس"},
    {"fa": "زیدوودین", "en": "Zidovudine", "type": "ضد ویروس"},
    {"fa": "مینرال‌میکس", "en": "Mineral Mix", "type": "تقویتی"},
    {"fa": "پرمیکس ویتامینه", "en": "Vitamin Premix", "type": "تقویتی"},
    {"fa": "ضد استرس", "en": "Anti-Stress Supplement", "type": "تقویتی"},
    {"fa": "آنتی‌اکسیدان", "en": "Antioxidant", "type": "تقویتی"},
    {"fa": "بتائین", "en": "Betaine", "type": "تقویتی"},
    {"fa": "اینوزیتول", "en": "Inositol", "type": "تقویتی"},
    {"fa": "اسپیرولینا", "en": "Spirulina", "type": "تقویتی"},
    {"fa": "عصاره آویشن", "en": "Thyme Extract", "type": "تقویتی"},
    {"fa": "اسانس اکالیپتوس", "en": "Eucalyptus Oil", "type": "تقویتی"},
    {"fa": "عصاره سیر", "en": "Garlic Extract", "type": "تقویتی"},
    {"fa": "سرکه سیب", "en": "Apple Cider Vinegar", "type": "تقویتی"},
    {"fa": "اسید بوتیریک", "en": "Butyric Acid", "type": "تقویتی"},
    {"fa": "اسید استیک", "en": "Acetic Acid", "type": "تقویتی"},
    {"fa": "فرمالین", "en": "Formalin", "type": "ضد عفونی"},
    {"fa": "گلوتارآلدهید", "en": "Glutaraldehyde", "type": "ضد عفونی"},
    {"fa": "هیپوکلریت سدیم", "en": "Sodium Hypochlorite", "type": "ضد عفونی"},
    {"fa": "بنزالکونیوم کلراید", "en": "Benzalkonium Chloride", "type": "ضد عفونی"},
    {"fa": "فنل", "en": "Phenol", "type": "ضد عفونی"},
    {"fa": "کرزول", "en": "Cresol", "type": "ضد عفونی"},
    {"fa": "ید", "en": "Iodine Solution", "type": "ضد عفونی"},
    {"fa": "سوخته آهک", "en": "Calcium Oxide", "type": "ضد عفونی"},
    {"fa": "پراستیک اسید", "en": "Peracetic Acid", "type": "ضد عفونی"},
    {"fa": "کلرین دی‌اکساید", "en": "Chlorine Dioxide", "type": "ضد عفونی"},
    {"fa": "آئروسیل", "en": "Aerosil", "type": "متفرقه"},
    {"fa": "کائولین پکتین", "en": "Kaolin Pectin", "type": "گوارشی"},
    {"fa": "نئواسکوایلن", "en": "Neo-Squilane", "type": "متفرقه"},
    {"fa": "روغن پارافین", "en": "Paraffin Oil", "type": "گوارشی"},
    {"fa": "روغن کرچک", "en": "Castor Oil", "type": "گوارشی"},
    {"fa": "میلبمایسین اکسیم", "en": "Milbemycin Oxime", "type": "ضد انگل"},
    {"fa": "فنبوکونازول", "en": "Fenbuconazole", "type": "ضد قارچ"},
    {"fa": "اکسی‌کوئینولین", "en": "Oxyquinoline", "type": "ضد عفونی"},
    {"fa": "کوپر سولفات", "en": "Copper Sulfate", "type": "ضد عفونی"},
    {"fa": "لینسید", "en": "Linseed", "type": "تقویتی"},
    {"fa": "زینک اکساید", "en": "Zinc Oxide", "type": "پوستی"},
    {"fa": "گوگرد", "en": "Sulfur", "type": "پوستی"},
    {"fa": "دی‌اتیل‌کاربامازین", "en": "Diethylcarbamazine", "type": "ضد انگل"},
    {"fa": "اکسیبندازول", "en": "Oxibendazole", "type": "ضد انگل"},
    {"fa": "کامبندازول", "en": "Cambendazole", "type": "ضد انگل"},
    {"fa": "بیتیونول", "en": "Bithionol", "type": "ضد انگل"},
    {"fa": "نیتروکسینیل", "en": "Nitroxinil", "type": "ضد انگل"},
    {"fa": "تایاپروست", "en": "Tiaprost", "type": "هورمونی"},
    {"fa": "لوپروستول", "en": "Luprostiol", "type": "هورمونی"},
    {"fa": "فنپروستالن", "en": "Fenprostalene", "type": "هورمونی"},
    {"fa": "آیوداید سدیم", "en": "Sodium Iodide", "type": "تقویتی"},
    {"fa": "نورفلورازون", "en": "Norflurazon", "type": "متفرقه"},
    {"fa": "فلوبندازول", "en": "Flubendazole", "type": "ضد انگل"},
    {"fa": "تترامیزول", "en": "Tetramisole", "type": "ضد انگل"},
    {"fa": "اپری‌مکتین", "en": "Eprinomectin", "type": "ضد انگل"},
    {"fa": "مونپانتل", "en": "Monepantel", "type": "ضد انگل"},
    {"fa": "درکوانتل", "en": "Derquantel", "type": "ضد انگل"},
    {"fa": "ابامکتین", "en": "Abamectin", "type": "ضد انگل"},
    {"fa": "کلوسوریل", "en": "Closulon", "type": "ضد انگل"},
    {"fa": "آمونیوم مولیبدات", "en": "Ammonium Molybdate", "type": "آنتی‌دوت"},
    {"fa": "کلسیم ای‌دی‌تی‌ای", "en": "Calcium EDTA", "type": "آنتی‌دوت"},
    {"fa": "متیلن بلو", "en": "Methylene Blue", "type": "آنتی‌دوت"},
    {"fa": "پلاسمای خون", "en": "Blood Plasma", "type": "سرم و مایعات"},
    {"fa": "خون کامل", "en": "Whole Blood", "type": "سرم و مایعات"},
    {"fa": "ژلاتین ساکسینات", "en": "Gelatin Succinate", "type": "سرم و مایعات"},
    {"fa": "پنی‌سیلین بنزاتین", "en": "Benzathine Penicillin", "type": "آنتی‌بیوتیک"},
    {"fa": "پنی‌سیلین پروکائین", "en": "Procaine Penicillin", "type": "آنتی‌بیوتیک"},
    {"fa": "تیلواکوزین", "en": "Tylvalosin", "type": "آنتی‌بیوتیک"},
    {"fa": "گاماپنتین", "en": "Gamithromycin", "type": "آنتی‌بیوتیک"},
    {"fa": "تولفنامیک اسید", "en": "Tolfenamic Acid", "type": "ضد التهاب"},
    {"fa": "ونلافاکسین", "en": "Venlafaxine", "type": "رفتاری"},
    {"fa": "میرتازاپین", "en": "Mirtazapine", "type": "رفتاری"},
    {"fa": "سوکسینیل‌کولین", "en": "Succinylcholine", "type": "بیهوشی"},
    {"fa": "اتومیدات", "en": "Etomidate", "type": "بیهوشی"},
    {"fa": "پانکرونیوم", "en": "Pancuronium", "type": "بیهوشی"},
]

DRUG_TYPES = ['آنتی‌بیوتیک', 'ضد انگل', 'ضد التهاب', 'کورتیکواستروئید', 'ضد حساسیت',
              'بیهوشی', 'ضد درد', 'ضد قارچ', 'تقویتی', 'سرم و مایعات',
              'قلبی‌عروقی', 'گوارشی', 'ضد سم', 'آنتی‌دوت', 'اورژانسی', 'هورمونی', 'سایر']


ORIGINS = ['گرمسیری', 'کبوترسانان', 'قرنطینه', 'گنجشکسانان', 'فاز۱', 'تکثیر فاز ۱',
           'سالن پرورش', 'شکاری فاز۱', 'فاز ۲', 'تکثیر فاز ۲', 'سالن مرغ خروس‌ها']

DRUG_UNITS = ['عدد', 'ورق', 'بسته', 'شیشه', 'ویال', 'آمپول', 'شربت',
              'میلی‌لیتر', 'لیتر', 'گرم', 'کیلوگرم', 'سرم', 'تیوب', 'ساشه', 'سی‌سی']

# ══════════════════════════════════════
# MODELS
# ══════════════════════════════════════

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active_user = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.String(20), default=shamsi_now)

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)

class Species(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    category = db.Column(db.String(50), default='زینتی')

class Bird(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(10), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    species_id = db.Column(db.Integer, db.ForeignKey('species.id'))
    breed = db.Column(db.String(100))
    sex = db.Column(db.String(20))
    age = db.Column(db.String(50))
    ring = db.Column(db.String(50))
    section = db.Column(db.String(100))
    weight = db.Column(db.String(20))
    status = db.Column(db.String(30), default='حاضر')
    treatment_status = db.Column(db.String(30), default='بدون درمان')
    entry_date = db.Column(db.String(20), default=shamsi_now)
    origin = db.Column(db.String(100))
    photo = db.Column(db.String(200))
    quantity = db.Column(db.Integer, default=1)
    sub_category = db.Column(db.String(50))
    initial_treatment = db.Column(db.Text)
    initial_drugs = db.Column(db.Text)
    needs_surgery = db.Column(db.Boolean, default=False)
    had_surgery = db.Column(db.Boolean, default=False)
    surgery_notes = db.Column(db.Text)
    discharge_reason = db.Column(db.Text)
    death_reason = db.Column(db.Text)
    discharge_date = db.Column(db.String(20))
    death_date = db.Column(db.String(20))
    created_at = db.Column(db.String(20), default=shamsi_now)

    species = db.relationship('Species', backref='birds')
    medical_records = db.relationship('MedicalRecord', backref='bird', lazy=True, order_by='MedicalRecord.id.desc()', cascade='all, delete-orphan')
    medications = db.relationship('Medication', backref='bird', lazy=True, order_by='Medication.id.desc()', cascade='all, delete-orphan')
    vaccines = db.relationship('Vaccine', backref='bird', lazy=True, order_by='Vaccine.id.desc()', cascade='all, delete-orphan')
    labs = db.relationship('LabResult', backref='bird', lazy=True, order_by='LabResult.id.desc()', cascade='all, delete-orphan')
    notes = db.relationship('Note', backref='bird', lazy=True, order_by='Note.id.desc()', cascade='all, delete-orphan')
    treatment_logs = db.relationship('TreatmentLog', backref='bird', lazy=True, order_by='TreatmentLog.id.desc()', cascade='all, delete-orphan')
    surgeries = db.relationship('SurgeryRecord', backref='bird', lazy=True, order_by='SurgeryRecord.id.desc()', cascade='all, delete-orphan')

class MedicalRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=False)
    date = db.Column(db.String(20), default=shamsi_now)
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    complaint = db.Column(db.Text)
    exam = db.Column(db.Text)
    diagnosis = db.Column(db.Text)
    plan = db.Column(db.Text)
    record_type = db.Column(db.String(30), default='visit')
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    vet = db.relationship('User')

class Medication(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=False)
    drug = db.Column(db.String(120), nullable=False)
    drug_en = db.Column(db.String(120))
    volume = db.Column(db.String(50))
    route = db.Column(db.String(50))
    start_date = db.Column(db.String(20))
    status = db.Column(db.String(20), default='فعال')
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    vet = db.relationship('User')

class MedInventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    drug_name = db.Column(db.String(120), nullable=False)
    drug_name_en = db.Column(db.String(120))
    drug_type = db.Column(db.String(50))
    concentration = db.Column(db.String(50))
    quantity = db.Column(db.Float, default=0)
    unit = db.Column(db.String(30))
    min_stock = db.Column(db.Float, default=0)
    notes = db.Column(db.Text)
    status = db.Column(db.String(20), default='موجود')

class NeedItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    drug_name = db.Column(db.String(120), nullable=False)
    drug_name_en = db.Column(db.String(120))
    drug_type = db.Column(db.String(50))
    quantity = db.Column(db.Float, default=0)
    unit = db.Column(db.String(30))
    notes = db.Column(db.Text)
    created_at = db.Column(db.String(30), default=shamsi_now_full)

class Vaccine(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=True)
    is_clinic = db.Column(db.Boolean, default=True)
    external_name = db.Column(db.String(120))
    species_name = db.Column(db.String(100))
    vaccine_name = db.Column(db.String(120), nullable=False)
    date = db.Column(db.String(20), default=shamsi_now)
    next_date = db.Column(db.String(20))
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    serial = db.Column(db.String(100))
    description = db.Column(db.Text)
    is_done = db.Column(db.Boolean, default=False)
    last_injection_date = db.Column(db.String(20))
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    vet = db.relationship('User')

class LabResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=True)
    is_clinic = db.Column(db.Boolean, default=True)
    external_name = db.Column(db.String(200))
    test_type = db.Column(db.String(100))
    date = db.Column(db.String(20), default=shamsi_now)
    lab_name = db.Column(db.String(120))
    result = db.Column(db.Text)
    interpretation = db.Column(db.Text)
    file_path = db.Column(db.String(200))
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    vet = db.relationship('User')

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=False)
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    vet = db.relationship('User')

class TreatmentLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=False)
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    medication_id = db.Column(db.Integer, db.ForeignKey('medication.id'))
    done_at = db.Column(db.String(30), default=shamsi_now_full)
    notes = db.Column(db.Text)
    vet = db.relationship('User')
    medication = db.relationship('Medication')

class SurgeryRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    bird_id = db.Column(db.Integer, db.ForeignKey('bird.id'), nullable=False)
    date = db.Column(db.String(20), default=shamsi_now)
    description = db.Column(db.Text)
    follow_up_date = db.Column(db.String(20))
    image = db.Column(db.String(200))
    vet_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    vet = db.relationship('User')

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    message = db.Column(db.Text, nullable=False)
    notif_type = db.Column(db.String(30))
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    user = db.relationship('User')

class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    action = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.String(30), default=shamsi_now_full)
    user = db.relationship('User')

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# ══════════════════════════════════════
# HELPERS
# ══════════════════════════════════════

def generate_bird_code():
    used_codes = {b.code for b in Bird.query.all()}
    num = 1
    while True:
        code = str(num).zfill(3)
        if code not in used_codes:
            return code
        num += 1

def notify_all_except(sender_id, message, notif_type='update'):
    users = User.query.filter(User.id != sender_id, User.is_active_user == True).all()
    for u in users:
        db.session.add(Notification(user_id=u.id, message=message, notif_type=notif_type))
    db.session.commit()

def log_activity(action):
    db.session.add(ActivityLog(user_id=current_user.id, action=action))
    db.session.commit()

def get_unread_count():
    if current_user.is_authenticated:
        return Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return 0

def get_daily_summary():
    today = shamsi_now()
    visits = MedicalRecord.query.filter(MedicalRecord.date == today).count()
    meds_given = TreatmentLog.query.filter(TreatmentLog.done_at.contains(today.replace('/', '')[:8]) | TreatmentLog.done_at.contains(today)).count()
    vaccines_today = Vaccine.query.filter(Vaccine.date == today).count()
    admissions = Bird.query.filter(Bird.entry_date == today).count()
    return {'visits': visits, 'meds': meds_given, 'vaccines': vaccines_today, 'admissions': admissions}

@app.context_processor
def inject_globals():
    uc = get_unread_count() if current_user.is_authenticated else 0
    return {'unread_count': uc, 'shamsi_now': shamsi_now(), 'shamsi_now_full': shamsi_now_full(), 'shamsi_time': shamsi_time()}

# API for drug autocomplete
@app.route('/api/drugs')
@login_required
def api_drugs():
    q = request.args.get('q', '').lower()
    results = [d for d in VET_DRUGS if q in d['fa'] or q in d['en'].lower()]
    return {'drugs': results[:20]}

@app.route('/health')
def health_check():
    return {'status': 'ok'}, 200

# ══════════════════════════════════════
# AUTH
# ══════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username', '').strip()).first()
        if user and user.check_password(request.form.get('password', '')):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('نام کاربری یا رمز عبور اشتباه است', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ══════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════

@app.route('/')
@login_required
def dashboard():
    present_birds = Bird.query.filter_by(status='حاضر').all()
    ornamental = sum((b.quantity or 1) for b in present_birds if b.species and b.species.category == 'زینتی')
    hunters = sum((b.quantity or 1) for b in present_birds if b.species and b.species.category == 'شکاری')
    domestic = sum((b.quantity or 1) for b in present_birds if b.species and b.species.category == 'اهلی')
    others = sum((b.quantity or 1) for b in present_birds if b.species and b.species.category == 'سایر')
    total = ornamental + hunters + domestic + others
    surgery_count = sum(1 for b in present_birds if b.needs_surgery)
    species_counts = db.session.query(Species.name, db.func.count(Bird.id)).join(Bird).filter(Bird.status == 'حاضر').group_by(Species.name).all()
    # Origin distribution
    from collections import defaultdict
    origin_data = defaultdict(lambda: {'cases': 0, 'total': 0})
    for b in present_birds:
        origin = b.origin or 'نامشخص'
        origin_data[origin]['cases'] += 1
        origin_data[origin]['total'] += (b.quantity or 1)
    origin_counts = sorted(origin_data.items(), key=lambda x: x[1]['total'], reverse=True)
    max_origin = max((d['total'] for _, d in origin_counts), default=1)
    upcoming_vaccines = Vaccine.query.filter(Vaccine.next_date != None, Vaccine.next_date != '', Vaccine.is_done == False).order_by(Vaccine.next_date).limit(5).all()
    upcoming_surgeries = SurgeryRecord.query.filter(SurgeryRecord.follow_up_date != None, SurgeryRecord.follow_up_date != '').order_by(SurgeryRecord.follow_up_date).limit(5).all()
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.id.desc()).limit(10).all()
    low_stock = MedInventory.query.filter(MedInventory.quantity <= 0).all()
    for item in low_stock:
        exists = Notification.query.filter(Notification.message.contains(item.drug_name), Notification.created_at.contains(shamsi_now())).first()
        if not exists:
            for u in User.query.filter_by(is_active_user=True).all():
                db.session.add(Notification(user_id=u.id, message=f'⚠️ داروی {item.drug_name} رو به اتمام است ({item.quantity} {item.unit}) — برای شارژ اقدام کنید', notif_type='alert'))
            db.session.commit()
    need_items = NeedItem.query.order_by(NeedItem.id.desc()).all()
    return render_template('dashboard.html', total=total, hunters=hunters, ornamental=ornamental,
        domestic=domestic, others=others, surgery_count=surgery_count, species_counts=species_counts,
        origin_counts=origin_counts, max_origin=max_origin,
        upcoming_vaccines=upcoming_vaccines,
        upcoming_surgeries=upcoming_surgeries, notifications=notifications, low_stock=low_stock,
        need_items=need_items)

# ══════════════════════════════════════
# NOTIFICATIONS
# ══════════════════════════════════════

@app.route('/notifications')
@login_required
def notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.id.desc()).limit(50).all()
    return render_template('notifications.html', notifications=notifs)

@app.route('/notifications/read-all', methods=['POST'])
@login_required
def read_all_notifications():
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return redirect(request.referrer or url_for('dashboard'))

# ══════════════════════════════════════
# BIRDS
# ══════════════════════════════════════

@app.route('/birds')
@login_required
def birds():
    tab = request.args.get('tab', 'present')
    search = request.args.get('search', '')
    status_map = {'present': 'حاضر', 'discharged': 'ترخیص', 'deceased': 'تلفات'}
    query = Bird.query.filter_by(status=status_map.get(tab, 'حاضر'))
    if search:
        query = query.filter(db.or_(Bird.name.contains(search), Bird.code.contains(search), Bird.ring.contains(search)))
    birds_list = query.order_by(Bird.code).all()
    all_species = Species.query.order_by(Species.name).all()
    return render_template('birds.html', birds=birds_list, tab=tab, search=search, all_species=all_species)

@app.route('/birds/new', methods=['GET', 'POST'])
@login_required
def new_bird():
    if request.method == 'POST':
        bird = Bird(
            code=request.form.get('code', generate_bird_code()).strip(),
            name=request.form.get('name', '').strip(),
            species_id=request.form.get('species_id'),
            breed=request.form.get('breed', '').strip(),
            sex=request.form.get('sex', ''),
            age=request.form.get('age', '').strip(),
            ring=request.form.get('ring', '').strip(),
            section=request.form.get('section', '').strip(),
            weight=request.form.get('weight', '').strip(),
            origin=request.form.get('origin', ''),
            entry_date=request.form.get('entry_date', shamsi_now()),
            initial_treatment=request.form.get('initial_treatment', '').strip(),
            initial_drugs=request.form.get('initial_drugs', '').strip(),
            treatment_status=request.form.get('treatment_status', 'بدون درمان'),
            needs_surgery=bool(request.form.get('needs_surgery')),
            surgery_notes=request.form.get('surgery_notes', '').strip(),
            quantity=int(request.form.get('quantity', 1) or 1),
            sub_category=request.form.get('sub_category', '').strip(),
        )
        if 'photo' in request.files:
            f = request.files['photo']
            if f and f.filename and allowed_file(f.filename):
                fn = f"bird_{bird.code}_{secure_filename(f.filename)}"
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                bird.photo = fn
        db.session.add(bird)
        db.session.commit()
        log_activity(f'پذیرش پرنده {bird.name} (کد {bird.code})')
        notify_all_except(current_user.id, f'پرنده جدید: {bird.name} (کد {bird.code}) — {current_user.full_name}', 'alert')
        flash(f'پرنده {bird.name} با کد {bird.code} ثبت شد', 'success')
        return redirect(url_for('bird_detail', bird_id=bird.id))
    all_species = Species.query.order_by(Species.name).all()
    return render_template('bird_form.html', bird=None, all_species=all_species, next_code=generate_bird_code(), origins=ORIGINS)

@app.route('/birds/<int:bird_id>')
@login_required
def bird_detail(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    tab = request.args.get('tab', 'info')
    return render_template('bird_detail.html', bird=bird, tab=tab)

@app.route('/birds/<int:bird_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_bird(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    if request.method == 'POST':
        new_code = request.form.get('code', bird.code).strip()
        if new_code != bird.code:
            existing = Bird.query.filter_by(code=new_code).first()
            if existing:
                flash('این کد قبلاً استفاده شده', 'error')
                return redirect(url_for('edit_bird', bird_id=bird.id))
            bird.code = new_code
        bird.name = request.form.get('name', bird.name).strip()
        bird.species_id = request.form.get('species_id', bird.species_id)
        bird.breed = request.form.get('breed', '').strip()
        bird.sex = request.form.get('sex', '')
        bird.age = request.form.get('age', '').strip()
        bird.ring = request.form.get('ring', '').strip()
        bird.section = request.form.get('section', '').strip()
        bird.weight = request.form.get('weight', '').strip()
        bird.origin = request.form.get('origin', '')
        bird.treatment_status = request.form.get('treatment_status', bird.treatment_status)
        bird.needs_surgery = bool(request.form.get('needs_surgery'))
        bird.surgery_notes = request.form.get('surgery_notes', '').strip()
        bird.initial_treatment = request.form.get('initial_treatment', '').strip()
        bird.initial_drugs = request.form.get('initial_drugs', '').strip()
        bird.quantity = int(request.form.get('quantity', bird.quantity or 1) or 1)
        bird.sub_category = request.form.get('sub_category', '').strip()
        if 'photo' in request.files:
            f = request.files['photo']
            if f and f.filename and allowed_file(f.filename):
                fn = f"bird_{bird.code}_{secure_filename(f.filename)}"
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                bird.photo = fn
        db.session.commit()
        log_activity(f'ویرایش {bird.name} (کد {bird.code})')
        notify_all_except(current_user.id, f'ویرایش {bird.name} (کد {bird.code}) — {current_user.full_name}', 'update')
        flash('اطلاعات بروزرسانی شد', 'success')
        return redirect(url_for('bird_detail', bird_id=bird.id))
    all_species = Species.query.order_by(Species.name).all()
    return render_template('bird_form.html', bird=bird, all_species=all_species, next_code=bird.code, origins=ORIGINS)

@app.route('/birds/<int:bird_id>/status', methods=['POST'])
@login_required
def update_treatment_status(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    bird.treatment_status = request.form.get('treatment_status', bird.treatment_status)
    db.session.commit()
    return redirect(request.referrer or url_for('bird_detail', bird_id=bird.id))

@app.route('/birds/<int:bird_id>/discharge', methods=['POST'])
@login_required
def discharge_bird(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    transfer_count = int(request.form.get('transfer_count', bird.quantity or 1) or 1)
    reason = request.form.get('reason', '').strip()
    if transfer_count >= (bird.quantity or 1):
        bird.status = 'ترخیص'
        bird.discharge_reason = reason
        bird.discharge_date = shamsi_now()
    else:
        bird.quantity = (bird.quantity or 1) - transfer_count
        base_code = bird.code + '-D'
        counter = 1
        new_code = base_code + str(counter)
        while Bird.query.filter_by(code=new_code).first():
            counter += 1
            new_code = base_code + str(counter)
        new_bird = Bird(
            code=new_code,
            name=bird.name, species_id=bird.species_id, breed=bird.breed,
            sex=bird.sex, age=bird.age, ring=bird.ring, section=bird.section,
            weight=bird.weight, origin=bird.origin, entry_date=bird.entry_date,
            quantity=transfer_count, sub_category=bird.sub_category,
            status='ترخیص', discharge_reason=reason, discharge_date=shamsi_now(),
            treatment_status=bird.treatment_status)
        db.session.add(new_bird)
    db.session.commit()
    log_activity(f'ترخیص {transfer_count} عدد از {bird.name} (کد {bird.code})')
    notify_all_except(current_user.id, f'ترخیص: {bird.name} ({transfer_count} عدد) — {current_user.full_name}', 'update')
    flash(f'{transfer_count} عدد از {bird.name} ترخیص شد', 'success')
    return redirect(url_for('birds'))

@app.route('/birds/<int:bird_id>/death', methods=['POST'])
@login_required
def death_bird(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    transfer_count = int(request.form.get('transfer_count', bird.quantity or 1) or 1)
    reason = request.form.get('reason', '').strip()
    if transfer_count >= (bird.quantity or 1):
        bird.status = 'تلفات'
        bird.death_reason = reason
        bird.death_date = shamsi_now()
    else:
        bird.quantity = (bird.quantity or 1) - transfer_count
        base_code = bird.code + '-X'
        counter = 1
        new_code = base_code + str(counter)
        while Bird.query.filter_by(code=new_code).first():
            counter += 1
            new_code = base_code + str(counter)
        new_bird = Bird(
            code=new_code,
            name=bird.name, species_id=bird.species_id, breed=bird.breed,
            sex=bird.sex, age=bird.age, ring=bird.ring, section=bird.section,
            weight=bird.weight, origin=bird.origin, entry_date=bird.entry_date,
            quantity=transfer_count, sub_category=bird.sub_category,
            status='تلفات', death_reason=reason, death_date=shamsi_now(),
            treatment_status=bird.treatment_status)
        db.session.add(new_bird)
    db.session.commit()
    log_activity(f'تلفات {transfer_count} عدد از {bird.name} (کد {bird.code})')
    notify_all_except(current_user.id, f'تلفات: {bird.name} ({transfer_count} عدد) — {current_user.full_name}', 'alert')
    flash(f'{transfer_count} عدد از {bird.name} در تلفات ثبت شد', 'warning')
    return redirect(url_for('birds'))

@app.route('/birds/<int:bird_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_bird(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    name = bird.name
    code = bird.code
    db.session.delete(bird)
    db.session.commit()
    log_activity(f'حذف پرنده {name} (کد {code})')
    flash(f'پرنده {name} (کد {code}) حذف شد — کد آزاد شد', 'success')
    return redirect(url_for('birds'))

@app.route('/birds/<int:bird_id>/upload-photo', methods=['POST'])
@login_required
def upload_bird_photo(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    if 'photo' in request.files:
        f = request.files['photo']
        if f and f.filename and allowed_file(f.filename):
            fn = f"bird_{bird.code}_{secure_filename(f.filename)}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            bird.photo = fn
            db.session.commit()
            flash('عکس آپلود شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bird.id))

# ══════════════════════════════════════
# MEDICAL RECORDS
# ══════════════════════════════════════

@app.route('/birds/<int:bird_id>/record', methods=['POST'])
@login_required
def add_record(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    rec = MedicalRecord(bird_id=bird.id, vet_id=current_user.id,
        date=request.form.get('date', shamsi_now()),
        complaint=request.form.get('complaint', '').strip(),
        exam=request.form.get('exam', '').strip(),
        diagnosis=request.form.get('diagnosis', '').strip(),
        plan=request.form.get('plan', '').strip(),
        record_type=request.form.get('record_type', 'visit'))
    db.session.add(rec)
    db.session.commit()
    log_activity(f'ویزیت {bird.name} — {rec.diagnosis or ""}')
    notify_all_except(current_user.id, f'ویزیت {bird.name}: {rec.diagnosis or "بدون تشخیص"} — {current_user.full_name}', 'update')
    flash('پرونده ثبت شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bird.id, tab='records'))

@app.route('/record/<int:rec_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_record(rec_id):
    rec = MedicalRecord.query.get_or_404(rec_id)
    bid = rec.bird_id
    db.session.delete(rec)
    db.session.commit()
    flash('ویزیت حذف شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bid, tab='records'))

# ══════════════════════════════════════
# NOTES
# ══════════════════════════════════════

@app.route('/birds/<int:bird_id>/note', methods=['POST'])
@login_required
def add_note(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    content = request.form.get('content', '').strip()
    if content:
        db.session.add(Note(bird_id=bird.id, vet_id=current_user.id, content=content))
        db.session.commit()
        notify_all_except(current_user.id, f'نوت جدید برای {bird.name} — {current_user.full_name}', 'note')
        flash('نوت ثبت شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bird.id, tab='notes'))

@app.route('/note/<int:note_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_note(note_id):
    note = Note.query.get_or_404(note_id)
    bid = note.bird_id
    db.session.delete(note)
    db.session.commit()
    flash('نوت حذف شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bid, tab='notes'))

# ══════════════════════════════════════
# MEDICATIONS
# ══════════════════════════════════════

@app.route('/medications')
@login_required
def medications():
    tab = request.args.get('tab', 'inventory')
    page = request.args.get('page', 1, type=int)
    if tab == 'inventory':
        pagination = MedInventory.query.order_by(MedInventory.drug_name).paginate(page=page, per_page=50, error_out=False)
        return render_template('medications.html', tab=tab, pagination=pagination, drug_types=DRUG_TYPES, drug_units=DRUG_UNITS)
    else:
        needs = NeedItem.query.order_by(NeedItem.id.desc()).all()
        return render_template('medications.html', tab=tab, needs=needs, drug_types=DRUG_TYPES, drug_units=DRUG_UNITS)

@app.route('/birds/<int:bird_id>/medication', methods=['POST'])
@login_required
def add_medication(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    drugs = request.form.getlist('drug')
    drugs_en = request.form.getlist('drug_en')
    volumes = request.form.getlist('volume')
    routes = request.form.getlist('route')
    start_dates = request.form.getlist('start_date')
    count = 0
    for i in range(len(drugs)):
        d = drugs[i].strip() if i < len(drugs) else ''
        if not d:
            continue
        med = Medication(bird_id=bird.id, vet_id=current_user.id,
            drug=d,
            drug_en=drugs_en[i].strip() if i < len(drugs_en) else '',
            volume=volumes[i].strip() if i < len(volumes) else '',
            route=routes[i] if i < len(routes) else '',
            start_date=start_dates[i] if i < len(start_dates) else '',
            status='فعال')
        db.session.add(med)
        count += 1
    db.session.commit()
    log_activity(f'{count} دارو برای {bird.name}')
    notify_all_except(current_user.id, f'{count} دارو برای {bird.name} — {current_user.full_name}', 'med')
    flash(f'{count} دارو ثبت شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bird.id, tab='meds'))

@app.route('/medication/<int:med_id>/edit', methods=['POST'])
@login_required
def edit_medication(med_id):
    med = Medication.query.get_or_404(med_id)
    med.drug = request.form.get('drug', med.drug).strip()
    med.drug_en = request.form.get('drug_en', med.drug_en or '').strip()
    med.volume = request.form.get('volume', med.volume or '').strip()
    med.route = request.form.get('route', med.route)
    med.start_date = request.form.get('start_date', med.start_date)
    db.session.commit()
    flash('دارو ویرایش شد', 'success')
    return redirect(url_for('bird_detail', bird_id=med.bird_id, tab='meds'))

@app.route('/medication/<int:med_id>/complete', methods=['POST'])
@login_required
def complete_medication(med_id):
    med = Medication.query.get_or_404(med_id)
    med.status = 'تکمیل'
    db.session.commit()
    return redirect(request.referrer or url_for('medications'))

@app.route('/medication/<int:med_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_medication(med_id):
    med = Medication.query.get_or_404(med_id)
    bid = med.bird_id
    db.session.delete(med)
    db.session.commit()
    flash('دارو حذف شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bid, tab='meds'))

@app.route('/inventory/add', methods=['POST'])
@login_required
def add_inventory():
    item = MedInventory(
        drug_name=request.form.get('drug_name', '').strip(),
        drug_name_en=request.form.get('drug_name_en', '').strip(),
        drug_type=request.form.get('drug_type', ''),
        concentration=request.form.get('concentration', '').strip(),
        quantity=float(request.form.get('quantity', 0) or 0),
        unit=request.form.get('unit', ''),
        min_stock=float(request.form.get('min_stock', 0) or 0),
        notes=request.form.get('notes', '').strip())
    db.session.add(item)
    db.session.commit()
    flash('دارو به موجودی اضافه شد', 'success')
    return redirect(url_for('medications', tab='inventory'))

@app.route('/inventory/<int:item_id>/update', methods=['POST'])
@login_required
def update_inventory(item_id):
    item = MedInventory.query.get_or_404(item_id)
    item.quantity = float(request.form.get('quantity', item.quantity) or 0)
    item.notes = request.form.get('notes', item.notes)
    db.session.commit()
    flash('موجودی بروزرسانی شد', 'success')
    return redirect(url_for('medications', tab='inventory'))

@app.route('/inventory/<int:item_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_inventory(item_id):
    item = MedInventory.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('حذف شد', 'success')
    return redirect(url_for('medications', tab='inventory'))

@app.route('/need/add', methods=['POST'])
@login_required
def add_need():
    item = NeedItem(
        drug_name=request.form.get('drug_name', '').strip(),
        drug_name_en=request.form.get('drug_name_en', '').strip(),
        drug_type=request.form.get('drug_type', ''),
        quantity=float(request.form.get('quantity', 0) or 0),
        unit=request.form.get('unit', ''),
        notes=request.form.get('notes', '').strip())
    db.session.add(item)
    db.session.commit()
    flash('به لیست نیاز اضافه شد', 'success')
    return redirect(url_for('medications', tab='needs'))

@app.route('/need/<int:item_id>/purchased', methods=['POST'])
@login_required
def need_purchased(item_id):
    need = NeedItem.query.get_or_404(item_id)
    inv = MedInventory(drug_name=need.drug_name, drug_name_en=need.drug_name_en,
        drug_type=need.drug_type, quantity=need.quantity, unit=need.unit, notes=need.notes)
    db.session.add(inv)
    db.session.delete(need)
    db.session.commit()
    flash(f'{need.drug_name} تهیه شد و به موجودی اضافه شد', 'success')
    return redirect(url_for('medications', tab='needs'))

@app.route('/need/<int:item_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_need(item_id):
    item = NeedItem.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('حذف شد', 'success')
    return redirect(url_for('medications', tab='needs'))

# ══════════════════════════════════════
# TREATMENT TICK
# ══════════════════════════════════════

@app.route('/birds/<int:bird_id>/treatment-tick', methods=['POST'])
@login_required
def treatment_tick(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    med_ids = request.form.getlist('medication_id')
    # support single hidden input too
    if not med_ids:
        single = request.form.get('medication_id')
        med_ids = [single] if single else [None]
    notes = request.form.get('notes', '').strip()
    for mid in med_ids:
        if not mid:
            mid = None
        log = TreatmentLog(bird_id=bird.id, vet_id=current_user.id,
            medication_id=mid,
            notes=notes)
        db.session.add(log)
        if mid:
            med = Medication.query.get(mid)
            drug_name = med.drug if med else 'درمان'
            notify_all_except(current_user.id, f'✅ {drug_name} برای {bird.name} — {current_user.full_name}', 'med')
    db.session.commit()
    flash('درمان ثبت شد ✅', 'success')
    return redirect(request.referrer or url_for('birds'))

# ══════════════════════════════════════
# VACCINES
# ══════════════════════════════════════

@app.route('/vaccines')
@login_required
def vaccines():
    all_vaccines = Vaccine.query.order_by(Vaccine.id.desc()).all()
    all_birds = Bird.query.order_by(Bird.code).all()
    all_species = Species.query.order_by(Species.name).all()
    return render_template('vaccines.html', vaccines=all_vaccines, all_birds=all_birds, all_species=all_species, origins=ORIGINS)

@app.route('/vaccines/add', methods=['POST'])
@login_required
def add_vaccine_general():
    is_clinic = request.form.get('is_clinic') == 'true'
    v = Vaccine(
        is_clinic=is_clinic,
        bird_id=request.form.get('bird_id') if is_clinic else None,
        external_name=request.form.get('external_name', '').strip() if not is_clinic else None,
        species_name=request.form.get('species_name', '').strip(),
        vaccine_name=request.form.get('vaccine_name', '').strip(),
        date=f"{request.form.get('date_y','')}/{request.form.get('date_m','')}/{request.form.get('date_d','')}",
        next_date=f"{request.form.get('next_y','')}/{request.form.get('next_m','')}/{request.form.get('next_d','')}",
        vet_id=current_user.id,
        serial=request.form.get('serial', '').strip(),
        description=request.form.get('description', '').strip())
    db.session.add(v)
    db.session.commit()
    log_activity(f'واکسن {v.vaccine_name}')
    flash('واکسن ثبت شد', 'success')
    return redirect(url_for('vaccines'))

@app.route('/birds/<int:bird_id>/vaccine', methods=['POST'])
@login_required
def add_vaccine(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    v = Vaccine(bird_id=bird.id, is_clinic=True, vet_id=current_user.id,
        vaccine_name=request.form.get('vaccine_name', '').strip(),
        date=f"{request.form.get('date_y','')}/{request.form.get('date_m','')}/{request.form.get('date_d','')}",
        next_date=f"{request.form.get('next_y','')}/{request.form.get('next_m','')}/{request.form.get('next_d','')}",
        serial=request.form.get('serial', '').strip(),
        description=request.form.get('description', '').strip())
    db.session.add(v)
    db.session.commit()
    notify_all_except(current_user.id, f'واکسن {v.vaccine_name} برای {bird.name} — {current_user.full_name}', 'vaccine')
    flash('واکسن ثبت شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bird.id, tab='vaccines'))

@app.route('/vaccine/<int:v_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_vaccine(v_id):
    v = Vaccine.query.get_or_404(v_id)
    db.session.delete(v)
    db.session.commit()
    flash('واکسن حذف شد', 'success')
    return redirect(request.referrer or url_for('vaccines'))

@app.route('/vaccine/<int:v_id>/edit', methods=['POST'])
@login_required
def edit_vaccine(v_id):
    v = Vaccine.query.get_or_404(v_id)
    v.vaccine_name = request.form.get('vaccine_name', v.vaccine_name).strip()
    v.date = request.form.get('date', v.date)
    v.next_date = request.form.get('next_date', v.next_date)
    v.serial = request.form.get('serial', v.serial)
    v.description = request.form.get('description', v.description)
    v.species_name = request.form.get('species_name', v.species_name)
    db.session.commit()
    flash('واکسن ویرایش شد', 'success')
    return redirect(request.referrer or url_for('vaccines'))

@app.route('/vaccine/<int:v_id>/mark-done', methods=['POST'])
@login_required
def mark_vaccine_done(v_id):
    v = Vaccine.query.get_or_404(v_id)
    v.is_done = True
    v.last_injection_date = shamsi_now()
    db.session.commit()
    flash('واکسن انجام شد ✅', 'success')
    return redirect(request.referrer or url_for('vaccines'))

# ══════════════════════════════════════
# LABS
# ══════════════════════════════════════

@app.route('/labs')
@login_required
def labs():
    all_labs = LabResult.query.order_by(LabResult.id.desc()).all()
    all_birds = Bird.query.order_by(Bird.code).all()
    return render_template('labs.html', labs=all_labs, all_birds=all_birds, origins=ORIGINS)

@app.route('/labs/add', methods=['POST'])
@login_required
def add_lab_general():
    is_clinic = request.form.get('is_clinic') == 'true'
    lab = LabResult(
        is_clinic=is_clinic,
        bird_id=request.form.get('bird_id') if is_clinic else None,
        external_name=request.form.get('external_name', '').strip() if not is_clinic else None,
        test_type=request.form.get('test_type', '').strip(),
        date=request.form.get('date', shamsi_now()),
        lab_name=request.form.get('lab_name', '').strip(),
        result=request.form.get('result', '').strip(),
        interpretation=request.form.get('interpretation', '').strip(),
        vet_id=current_user.id)
    if 'file' in request.files:
        f = request.files['file']
        if f and f.filename and allowed_file(f.filename):
            fn = f"lab_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(f.filename)}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            lab.file_path = fn
    db.session.add(lab)
    db.session.commit()
    log_activity(f'آزمایش {lab.test_type}')
    flash('آزمایش ثبت شد', 'success')
    return redirect(url_for('labs'))

@app.route('/birds/<int:bird_id>/lab', methods=['POST'])
@login_required
def add_lab(bird_id):
    bird = Bird.query.get_or_404(bird_id)
    lab = LabResult(bird_id=bird.id, is_clinic=True, vet_id=current_user.id,
        test_type=request.form.get('test_type', '').strip(),
        date=request.form.get('date', shamsi_now()),
        lab_name=request.form.get('lab_name', '').strip(),
        result=request.form.get('result', '').strip(),
        interpretation=request.form.get('interpretation', '').strip())
    if 'file' in request.files:
        f = request.files['file']
        if f and f.filename and allowed_file(f.filename):
            fn = f"lab_{bird.code}_{secure_filename(f.filename)}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            lab.file_path = fn
    db.session.add(lab)
    db.session.commit()
    notify_all_except(current_user.id, f'آزمایش {lab.test_type} برای {bird.name} — {current_user.full_name}', 'lab')
    flash('آزمایش ثبت شد', 'success')
    return redirect(url_for('bird_detail', bird_id=bird.id, tab='labs'))

@app.route('/lab/<int:lab_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_lab(lab_id):
    lab = LabResult.query.get_or_404(lab_id)
    db.session.delete(lab)
    db.session.commit()
    flash('آزمایش حذف شد', 'success')
    return redirect(request.referrer or url_for('labs'))

@app.route('/lab/<int:lab_id>/edit', methods=['POST'])
@login_required
def edit_lab(lab_id):
    lab = LabResult.query.get_or_404(lab_id)
    lab.test_type = request.form.get('test_type', lab.test_type).strip()
    lab.date = request.form.get('date', lab.date)
    lab.lab_name = request.form.get('lab_name', lab.lab_name)
    lab.result = request.form.get('result', lab.result)
    lab.interpretation = request.form.get('interpretation', lab.interpretation)
    if 'file' in request.files:
        f = request.files['file']
        if f and f.filename and allowed_file(f.filename):
            fn = f"lab_edit_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(f.filename)}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            lab.file_path = fn
    db.session.commit()
    flash('آزمایش ویرایش شد', 'success')
    return redirect(request.referrer or url_for('labs'))

# ══════════════════════════════════════
# SURGERY
# ══════════════════════════════════════

@app.route('/surgery')
@login_required
def surgery():
    operated = SurgeryRecord.query.order_by(SurgeryRecord.id.desc()).all()
    needs_surgery = Bird.query.filter_by(needs_surgery=True, status='حاضر').all()
    all_birds = Bird.query.filter_by(status='حاضر').order_by(Bird.code).all()
    return render_template('surgery.html', operated=operated, needs_surgery=needs_surgery, all_birds=all_birds)

@app.route('/surgery/add', methods=['POST'])
@login_required
def add_surgery():
    bird_id = request.form.get('bird_id')
    bird = Bird.query.get_or_404(bird_id)
    sr = SurgeryRecord(bird_id=bird.id, vet_id=current_user.id,
        date=request.form.get('date', shamsi_now()),
        description=request.form.get('description', '').strip(),
        follow_up_date=request.form.get('follow_up_date', '').strip())
    if 'image' in request.files:
        f = request.files['image']
        if f and f.filename and allowed_file(f.filename):
            fn = f"surgery_{bird.code}_{secure_filename(f.filename)}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            sr.image = fn
    bird.had_surgery = True
    bird.needs_surgery = False
    db.session.add(sr)
    db.session.commit()
    log_activity(f'جراحی {bird.name} (کد {bird.code})')
    notify_all_except(current_user.id, f'جراحی: {bird.name} (کد {bird.code}) — {current_user.full_name}', 'alert')
    flash('جراحی ثبت شد', 'success')
    return redirect(url_for('surgery'))

@app.route('/surgery/<int:sr_id>/edit', methods=['POST'])
@login_required
def edit_surgery(sr_id):
    sr = SurgeryRecord.query.get_or_404(sr_id)
    sr.date = request.form.get('date', sr.date)
    sr.description = request.form.get('description', sr.description).strip()
    sr.follow_up_date = request.form.get('follow_up_date', sr.follow_up_date).strip()
    if 'image' in request.files:
        f = request.files['image']
        if f and f.filename and allowed_file(f.filename):
            fn = f"surgery_{sr.bird.code}_{secure_filename(f.filename)}"
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            sr.image = fn
    db.session.commit()
    flash('جراحی بروزرسانی شد', 'success')
    return redirect(url_for('surgery'))

@app.route('/surgery/<int:sr_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_surgery(sr_id):
    sr = SurgeryRecord.query.get_or_404(sr_id)
    db.session.delete(sr)
    db.session.commit()
    flash('جراحی حذف شد', 'success')
    return redirect(url_for('surgery'))

# ══════════════════════════════════════
# REPORTS
# ══════════════════════════════════════

@app.route('/reports')
@login_required
def reports():
    period = request.args.get('period', 'weekly')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    month = request.args.get('month', '')
    year = request.args.get('year', str(jdatetime.date.today().year))
    shamsi_months = ['فروردین','اردیبهشت','خرداد','تیر','مرداد','شهریور','مهر','آبان','آذر','دی','بهمن','اسفند']
    return render_template('reports.html', period=period, date_from=date_from, date_to=date_to, month=month, year=year,
        shamsi_months=shamsi_months,
        birds_present=Bird.query.filter_by(status='حاضر').count(),
        birds_discharged=Bird.query.filter_by(status='ترخیص').count(),
        birds_deceased=Bird.query.filter_by(status='تلفات').count(),
        surgery_count=SurgeryRecord.query.count())

@app.route('/reports/export/excel')
@login_required
def export_report():
    from openpyxl import Workbook
    period = request.args.get('period', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    month = request.args.get('month', '')
    year = request.args.get('year', str(jdatetime.date.today().year))

    def in_range(date_str):
        if not date_str:
            return True
        if date_from and date_to:
            return date_from <= date_str <= date_to
        if month and year:
            m = int(month)
            prefix = f"{year}/{m:02d}/"
            return date_str.startswith(prefix)
        return True

    wb = Workbook()
    ws = wb.active
    ws.title = 'پذیرش‌شده‌ها'
    ws.sheet_view.rightToLeft = True
    ws.append(['کد', 'نام', 'گونه', 'جنسیت', 'وزن', 'بخش', 'منشا', 'تعداد', 'تاریخ ورود', 'وضعیت'])
    for b in Bird.query.filter_by(status='حاضر').all():
        if in_range(b.entry_date):
            ws.append([b.code, b.name, b.species.name if b.species else '', b.sex, b.weight, b.section, b.origin or '', b.quantity or 1, b.entry_date, b.treatment_status])

    ws2 = wb.create_sheet('ترخیص‌شده')
    ws2.sheet_view.rightToLeft = True
    ws2.append(['کد', 'نام', 'گونه', 'تاریخ ترخیص', 'دلیل بهبود'])
    for b in Bird.query.filter_by(status='ترخیص').all():
        if in_range(b.discharge_date):
            ws2.append([b.code, b.name, b.species.name if b.species else '', b.discharge_date, b.discharge_reason])

    ws3 = wb.create_sheet('تلفات')
    ws3.sheet_view.rightToLeft = True
    ws3.append(['کد', 'نام', 'گونه', 'تاریخ', 'علت'])
    for b in Bird.query.filter_by(status='تلفات').all():
        if in_range(b.death_date):
            ws3.append([b.code, b.name, b.species.name if b.species else '', b.death_date, b.death_reason])

    ws4 = wb.create_sheet('جراحی‌ها')
    ws4.sheet_view.rightToLeft = True
    ws4.append(['پرنده', 'کد', 'تاریخ جراحی', 'توضیحات', 'پیگیری', 'دامپزشک'])
    for s in SurgeryRecord.query.all():
        if in_range(s.date):
            ws4.append([s.bird.name, s.bird.code, s.date, s.description, s.follow_up_date, s.vet.full_name if s.vet else ''])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'report_{shamsi_now().replace("/","")}.xlsx',
                    as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════
# ADMIN
# ══════════════════════════════════════

@app.route('/admin')
@login_required
@admin_required
def admin_panel():
    users = User.query.all()
    species_list = Species.query.order_by(Species.name).all()
    return render_template('admin.html', users=users, species_list=species_list)

@app.route('/admin/user/add', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form.get('username', '').strip()
    if User.query.filter_by(username=username).first():
        flash('نام کاربری تکراری است', 'error')
        return redirect(url_for('admin_panel'))
    user = User(username=username, full_name=request.form.get('full_name', '').strip(),
                is_admin=bool(request.form.get('is_admin')))
    user.set_password(request.form.get('password', ''))
    db.session.add(user)
    db.session.commit()
    flash(f'کاربر {user.full_name} اضافه شد', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('نمی‌توانید خودتان را حذف کنید', 'error')
        return redirect(url_for('admin_panel'))
    db.session.delete(user)
    db.session.commit()
    flash('کاربر حذف شد', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/species/add', methods=['POST'])
@login_required
@admin_required
def add_species():
    name = request.form.get('name', '').strip()
    category = request.form.get('category', 'زینتی')
    if name:
        db.session.add(Species(name=name, category=category))
        db.session.commit()
        flash(f'گونه {name} اضافه شد', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/species/<int:sp_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_species(sp_id):
    sp = Species.query.get_or_404(sp_id)
    if sp.birds:
        flash('این گونه پرنده دارد و قابل حذف نیست', 'error')
    else:
        db.session.delete(sp)
        db.session.commit()
        flash('گونه حذف شد', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/backup')
@login_required
@admin_required
def backup_db():
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', 'bird_clinic.db')
    if os.path.exists(db_path):
        return send_file(db_path, download_name=f'backup_{shamsi_now().replace("/","")}.db', as_attachment=True)
    flash('فایل دیتابیس یافت نشد', 'error')
    return redirect(url_for('admin_panel'))

# ══════════════════════════════════════
# INIT
# ══════════════════════════════════════

def init_db():
    with app.app_context():
        db.create_all()
        # Migration: add missing columns safely
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        def has_column(table, column):
            try:
                cols = [c['name'] for c in inspector.get_columns(table)]
                return column in cols
            except Exception:
                return True
        migrations = [
            ('medication', 'volume', 'VARCHAR(50)'),
            ('bird', 'quantity', 'INTEGER DEFAULT 1'),
            ('bird', 'sub_category', 'VARCHAR(50)'),
            ('vaccine', 'is_done', 'BOOLEAN DEFAULT FALSE'),
            ('vaccine', 'last_injection_date', 'VARCHAR(20)'),
        ]
        for table, col, col_type in migrations:
            if not has_column(table, col):
                try:
                    db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {col} {col_type}'))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        if not User.query.first():
            admin = User(username='admin', full_name='دکتر مهدیار رمزگویان', is_admin=True)
            admin.set_password('admin123')
            db.session.add(admin)
            vet2 = User(username='vet1', full_name='دکتر احمدی', is_admin=False)
            vet2.set_password('vet123')
            db.session.add(vet2)
            db.session.commit()
            for name, cat in [('زینتی','زینتی'),('شکاری','شکاری'),('اهلی','اهلی'),('سایر','سایر')]:
                db.session.add(Species(name=name, category=cat))
            db.session.commit()
        # Ensure 4 default species exist for existing databases
        for name in ['زینتی', 'شکاری', 'اهلی', 'سایر']:
            if not Species.query.filter_by(name=name).first():
                db.session.add(Species(name=name, category=name))
        db.session.commit()

# ══════════════════════════════════════
# MIGRATION (one-time)
# ══════════════════════════════════════
@app.route('/migrate-db', methods=['GET', 'POST'])
@login_required
@admin_required
def migrate_db():
    if request.method == 'GET':
        return '''<!DOCTYPE html><html dir="rtl"><head><meta charset="utf-8"><title>انتقال دیتا</title>
        <style>body{font-family:sans-serif;max-width:500px;margin:50px auto;text-align:center;}
        input,button{margin:10px;padding:10px;}</style></head><body>
        <h2>📦 انتقال دیتا از SQLite</h2>
        <p>فایل bird_clinic.db رو آپلود کن</p>
        <form method="POST" enctype="multipart/form-data">
        <input type="file" name="db_file" required><br>
        <button type="submit" style="background:#16a34a;color:white;border:none;border-radius:8px;padding:12px 24px;cursor:pointer;">شروع انتقال</button>
        </form></body></html>'''
    import sqlite3, tempfile
    f = request.files.get('db_file')
    if not f:
        flash('فایلی انتخاب نشد', 'error')
        return redirect(url_for('migrate_db'))
    tmp_path = os.path.join('/tmp', 'migrate_bird_clinic.db')
    f.save(tmp_path)
    results = []
    try:
        conn = sqlite3.connect(tmp_path)
        conn.execute('SELECT count(*) FROM sqlite_master')
        conn.row_factory = sqlite3.Row
        tables_order = ['user','species','bird','medical_record','medication',
            'need_item','lab_result','surgery_record','activity_log',
            'notification','vaccine','treatment_log','note','med_inventory']
        from sqlalchemy import text, inspect
        inspector = inspect(db.engine)
        # Boolean columns that need casting
        bool_cols = {'is_admin','is_active_user','needs_surgery','had_surgery',
                     'is_clinic','is_read','is_done'}
        for table in tables_order:
            try:
                rows = conn.execute(f'SELECT * FROM "{table}"').fetchall()
                if not rows:
                    results.append(f'⏭️ {table}: empty')
                    continue
                sqlite_cols = [desc[0] for desc in conn.execute(f'SELECT * FROM "{table}" LIMIT 1').description]
                # Get PostgreSQL columns
                try:
                    pg_cols = [c['name'] for c in inspector.get_columns(table)]
                except:
                    pg_cols = sqlite_cols
                # Only use columns that exist in both
                valid_cols = [c for c in sqlite_cols if c in pg_cols]
                db.session.execute(text(f'TRUNCATE TABLE "{table}" CASCADE'))
                for row in rows:
                    raw = dict(zip(sqlite_cols, row))
                    values = {}
                    for c in valid_cols:
                        v = raw[c]
                        if c in bool_cols:
                            v = bool(v) if v is not None else False
                        values[c] = v
                    placeholders = ', '.join([f':{c}' for c in valid_cols])
                    col_names = ', '.join([f'"{c}"' for c in valid_cols])
                    db.session.execute(text(f'INSERT INTO "{table}" ({col_names}) VALUES ({placeholders})'), values)
                db.session.commit()
                try:
                    max_id = conn.execute(f'SELECT MAX(id) FROM "{table}"').fetchone()[0]
                    if max_id:
                        db.session.execute(text(f"SELECT setval(pg_get_serial_sequence('{table}', 'id'), {max_id})"))
                        db.session.commit()
                except:
                    db.session.rollback()
                results.append(f'✅ {table}: {len(rows)} rows')
            except Exception as e:
                db.session.rollback()
                results.append(f'⚠️ {table}: {e}')
        conn.close()
    except Exception as e:
        results.append(f'❌ Error: {e}')
    finally:
        os.unlink(tmp_path)
    return '<br>'.join(results) + '<br><br><a href="/">برو به داشبورد</a>'

init_db()

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
